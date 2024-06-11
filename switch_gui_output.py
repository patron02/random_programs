# This program imports excel data and exports it onto a txt file
# Place the xlsx file in the same path as the py file but not in the python script folder
# You need to have openpyxl library installed


from typing_extensions import is_protocol
import openpyxl

global version, configuration, build_date, remote_username, remote_password, location, password_user, network, vlans_list, network_management_vlan, rtu_cab, names_list, cli_prompt, password_user, max_temp, min_temp, max_humidity, min_humidity

version = "HiOS-2S-09.1.00"
configuration = "SWitch Name"
build_date = "2022-08-24 12:58"
remote_username = "admin"
remote_password = "123456"
cli_prompt = "123456"
password_admin = "123456"
password_user = "user"
network = "Thin Client Network"
vlans_list = "36,42,351,2016"  #EDIT
names_list = "trunk-36-net, disabled-42, tcn-351-net, mgmttcn-2016-net"
max_temp = 70
min_temp = 0
max_humidity = 95
min_humidity = 5

#EDIT"
network_management_vlan = "2016"  #EDIT
rtu_cab = "0RTU-123-1234"  #EDIT

from openpyxl import Workbook, load_workbook

wb = load_workbook('VLAN_database.xlsx', data_only=True)  #EDIT
ws = wb["TCN"]  #EDIT
ws_vlandatabase = wb["VLAN Database"]

# Open a separate txt file for each switch
for row in range(4, ws.max_row + 1):
    global tag
    tag = ws.cell(row=row, column=2).value
    filename = f"{tag}.txt"

    prev_tag = ws.cell(row=row - 1, column=2).value
    next_tag = ws.cell(row=row + 1, column=2).value

    # Set up variables you want to export to txt
    tag = ws.cell(row=row, column=2).value
    location = ws.cell(row=row, column=5).value
    ip = ws.cell(row=row, column=11).value
    gateway = ip[:ip.rfind('.') + 1] + '1'
    l2 = ws.cell(row=row, column=13).value
    port = ws.cell(row=row, column=15).value
    state = ws.cell(row=row, column=17).value
    speed = ws.cell(row=row, column=18).value
    tagged = ws.cell(row=row, column=23).value
    vlantags = str(ws.cell(row=row, column=24).value)
    vlanpvid = ws.cell(row=row, column=25).value

    # Split up the array
    vlans = [vlan.strip() for vlan in vlans_list.split(',')]
    names = names_list.split(',')
    parts = [part.strip() for part in vlantags.split(',')]

    # Print to txt file (change the format if needed)
    if tag != prev_tag:
        with open(filename, 'w') as f:
            f.write(f"! Copyright Company Inc\n")
            f.write(f"! Project \n")
            f.write(f"! 3rd Party Company\n")
            f.write(f"! Cabinet Name {rtu_cab}\n")
            f.write(f"! {network}\n")
            f.write(f"! Build Date: {build_date}\n\n")

            f.write(f"network dhcp config-load disable\n")
            f.write(f"network parms {ip} 255.255.255.0 {gateway}\n\n")

            f.write(
                f"no network ipv6 operation\nnetwork ipv6 protocol none\n\n")

            f.write(f"network hidiscovery operation disable\n")
            f.write(f"network hidiscovery mode read-only\n\n")

            f.write(
                f"network management priority dot1p 7\nnetwork management access delete 1\nnetwork management access delete 1\n"
            )

            f.write(
                f"network management access add 1 ip 192.168.216.0 mask 24 http disable https enable snmp enable telnet disable iec61850-mms disable modbus-tcp disable ssh enable ethernet-ip disable profinet-io disable\nnetwork management access status 1 enable\nnetwork management access add 2 ip 192.168.241.0 mask 24 http disable https enable snmp enable telnet disable iec61850-mms disable modbus-tcp disable ssh enable ethernet-ip disable profinet-io disable\nnetwork management access status 2 enable\nnetwork management access add 3 ip 172.16.4.0 mask 22 http disable https disable snmp enable telnet disable iec61850-mms disable modbus-tcp disable ssh disable ethernet-ip disable profinet-io disable\nnetwork management access status 3 enable network management access add 4 ip 172.17.4.0 mask 22 http disable https disable snmp enable telnet disable iec61850-mms disable modbus-tcp disable ssh disable ethernet-ip disable profinet-io disable\nnetwork management access status 4 enable\nnetwork management access operation\n\n"
            )

            f.write(f"cli prompt {cli_prompt}\n")
            f.write(f"cli serial-timeout 5\n")
            f.write(f"cli banner operation enable\n\n")

            f.write(f"configure\n")
            f.write(f"system contact Service_Engineer\n")
            f.write(f"system location {rtu_cab}\n")
            f.write(f"system name {tag}\n\n")

            f.write(f"config envm auto-update sd disable \n")
            f.write(f"config envm config-save sd disable \n")
            f.write(f"config envm sshkey-auto-update sd disable \n")
            f.write(f"config envm load-priority sd disable \n\n")

            f.write(f"clock summer-time mode usa\n\n")

            f.write(
                f"sntp client server add 1 172.16.4.1 port 123 description DC1\n"
            )
            f.write(
                f"sntp client server add 2 172.16.4.2 port 123 description DC2\n"
            )
            f.write(f"sntp client server mode 1 enable\n")
            f.write(f"sntp client server mode 2 enable\n")
            f.write(f"sntp client operation\n")

            f.write(f"passwords min-special-chars 0\n")
            f.write(f"users password admin {password_admin}\n")
            f.write(f"users add user\n")
            f.write(f"users password user {password_user}\n")
            f.write(f"users enable user\n\n")

            f.write(f"users password-policy-check admin enable\n")
            f.write(f"users password-policy-check user enable\n")
            f.write(f"passwords min-length 12\n")
            f.write(f"passwords min-special-chars 0\n\n")

            f.write(f"temperature upper-limit {max_temp}\n")
            f.write(f"temperature lower-limit {min_temp}\n\n")

            f.write(f"humidity upper-limit {max_humidity}\n")
            f.write(f"humidity lower-limit {min_humidity}\n\n")

            f.write(
                f"no telnet server\nno http server\nno snmp access version v1\nno snmp access version v2\nssh max-sessions 2\nssh timeout 5\n\n"
            )

            f.write(
                f"https fingerprint-type sha256\nhttps certificate delete\nhttps certificate generate\n\n"
            )

            f.write(f"cos-queue max-bandwidth 7 20\n\n")

            f.write(
                f"dos icmp-smurf-attack\ndos ip-land enable\ndos tcp-null\ndos tcp-syn-fin\ndos tcp-xmas\n\n"
            )

            f.write(
                f"mac notification interval 30\nmac notification operation\n\n"
            )

            f.write(
                f"system pre-login-banner operation\nsystem pre-login-banner text '***{network}***\nThis system is the property of 3rd Party Company\nOnly authorized personnel may use this system and all usage shall comply with the company security policy. Disconnect immediately if you do not fully agree on the company security policy or if the connection is unauthorized.\nUnauthorized use of this system is strictly prohibited and may be subject to criminal prosecution.\nAny usage of the system can be monitored and logged.\n\n'"
            )

            f.write(f"exit\nvlan database\n")
            for name, vlan in zip(names, vlans):
                f.write(f"vlan add {vlan}\n")
                f.write(f"name {vlan} {name}\n")
            f.write(f"exit\nconfigure\n\n")

    # This section writes the information for the individual ports
    with open(filename, 'a') as f:
        f.write(f"interface 1/{port}\n")

        if vlantags == "42":
            f.write(f"shutdown\n")
        else:
            f.write(f"no shutdown\n")

        f.write("mtu 1518\nutilization alarm-threshhold upper 7000\n")

        f.write(f"vlan participation auto 1\n")
        for part in parts:
            f.write(f"vlan particpation include {part}\n")
            if tagged != "U":
                for part in parts:
                    f.write(f"vlan tagging {part}\n")

        f.write(f"vlan pvid {vlanpvid}\n")
        # f.write(f"spanning-tree edge-port\nno device-status link-alarm\n")
        f.write(f"exit\n\n")

    if tag != next_tag:
        with open(filename, 'a') as f:

            f.write(
                f"device-status monitor link-failure\ndevice-status monitor power-supply 1\ndevice-status monitor power-supply 2\ndevice-status monitor temperature\ndevice-status monitor humidity\ndevice-status trap enable\ndevice-status monitor ring-redundancy\n\n"
            )

            f.write(
                f"signal-contact 1 mode manual\nsignal-contact 1 state close\nno signal-contact 1 monitor temperature\nno signal-contact 1 monitor power-supply 1\nno signal-contact 1 monitor power-supply 2\nno signal-contact 1 monitor humidity\n\n"
            )

            f.write(
                f"security-status monitor extnvm-load-unsecure\nsecurity-status monitor extnvm-upd-enabled\nsecurity-status monitor pwd-change\nsecurity-status monitor pwd-min-length\nsecurity-status monitor pwd-policy-config\nsecurity-status monitor pwd-policy-inactive\n\n"
            )

            f.write(
                f"security-status monitor hidisc-enabled\nsecurity-status monitor http-enabled\nsecurity-status monitor telnet-enabled\nsecurity-status monitor snmp-unsecure\n\n"
            )

            f.write(
                f"no security-status monitor https-certificate\nno security-status monitor no-link-enabled\nno security-status monitor sysmon-enabled\n\n"
            )

            f.write(f"security-statis trap enable\n\n")

            f.write(f"notice type systemlog\nnotice type audittrail\n\n")

            f.write(
                f"spanning-tree bpdu-guard\nauto-disable reason bpdu-rate enable\ninterface all\nauto-disable timer 30\nexit\n\n"
            )

            f.write(f"!Exit configure mode\nexit\n\n")
            f.write(f"network management vlan {network_management_vlan}\n")
            f.write(f"save\n\nreboot\n")

