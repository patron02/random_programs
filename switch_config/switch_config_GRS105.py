# -*- coding: utf-8 -*-
#!/usr/bin/python
# Title: Switch Configuration GRS105
#
# Project: 
#
# Co-Authors: Laura, Linda, Aline
#
# Program Revision: 02
#
# Date: June 2024
#
# Program Description:
# This program imports excel data and exports it onto a txt file
# Place the xlsx file in the same path as the py file but not in the python script folder
# You need to have openpyxl library installed
#
#
#
#
# Revision History
#--------------------------------------------------
# 001 | May 2024 | Laura | Initial Program
# 002 | June 2024 | Laura, Linda, Aline | Formatting
# 003 | July 2024 | Laura | Powershell GUI incorporation 
#
#
################################################################################################

import openpyxl
import datetime
import sys
import os

base_dir = "./Config Files"
if not os.path.exists(base_dir):
    os.makedirs(base_dir)

script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

global version, configuration, build_date, remote_username, remote_password
global password_user, max_temp, min_temp, max_humidity, min_humidity
global mtu, project_name, customer, network_tab

# Edit this section with the relevant information 

version = "HiOS-2S-09.1.00"
configuration = "GRS105/106"
project_name = sys.argv[1]
customer = sys.argv[2]
build_date = str(datetime.datetime.now())
remote_username = sys.argv[3]
remote_password = sys.argv[4]
password_admin = sys.argv[5]
password_user = sys.argv[6]
max_temp = sys.argv[8]
min_temp = sys.argv[7]
max_humidity = 95
min_humidity = 5
mtu = "1518"
network_tab = sys.argv[9] 
excel_file = str(sys.argv[10])

from openpyxl import Workbook, load_workbook


#wb = load_workbook(excel_file ,data_only=True) 
wb = load_workbook(excel_file ,data_only=True) # Replace with relevant file
ws = wb[network_tab] # Replace with relevant tab name 
ws_vlandatabase = wb["VLAN Database"]
ws_subnets = wb["Subnets and VLANs"]

# Find all the unique vlans for that network
vlans = set()
columns_to_search = ['X', 'Y']
for col in columns_to_search:
    column = ws[col]
    for cell in column:
        if cell.value:
            if isinstance(cell.value, str):
                numbers = cell.value.split(',')
            else:
                numbers = [str(cell.value)]  
            for number in numbers:
                try:
                    vlans.add(int(number.strip()))
                except ValueError:
                    pass  
vlans = sorted(vlans)

# Find the manager for this network
network_management_vlan = None
for number in vlans:
    for row in ws_vlandatabase.iter_rows(min_row=2, max_row=ws_vlandatabase.max_row, min_col=1, max_col=3):
        num_col_a = row[0].value
        num_col_b = row[1].value
        num_col_c = row[2].value
        if num_col_a == number:
            if ("mgmt" in str(num_col_b).lower() or "management" in str(num_col_b).lower() or "manager" in str(num_col_b).lower() or
                "mgmt" in str(num_col_c).lower() or "management" in str(num_col_c).lower() or "manager" in str(num_col_c).lower()):
                network_management_vlan = number
                break  
    if network_management_vlan:
        break

# Open a separate txt file for each switch
for row in range(4, ws.max_row + 1):
    global tag
    tag = ws.cell(row=row, column=2).value
    filename = f"{tag}.config"

    file_path = os.path.join(base_dir, filename)

    prev_tag = ws.cell(row=row - 1, column=2).value
    next_tag = ws.cell(row=row + 1, column=2).value

    # Set up variables you want to export to txt
    tag = ws.cell(row=row, column=2).value
    location = ws.cell(row=row, column=5).value
    ip = ws.cell(row=row, column=11).value
    gateway = ip[:ip.rfind('.') + 1] + '1'
    l2 = ws.cell(row=row, column=13).value
    rstp = str(ws.cell(row=row, column=14).value)
    port = ws.cell(row=row, column=15).value
    state = ws.cell(row=row, column=17).value
    speed = str(ws.cell(row=row, column=18).value)
    tagged = ws.cell(row=row, column=23).value
    vlantags = str(ws.cell(row=row, column=24).value)
    vlanpvid = ws.cell(row=row, column=25).value
    network = str(ws.cell(row=4, column=3).value)

    # Find corresponding name to Vlans 
    vlan_names = []
    for vlan in vlans:
        found = False 
        # Iterate over rows in VLAN Database to find the VLAN number
        for row in ws_vlandatabase.iter_rows(min_row=1, max_row=ws_vlandatabase.max_row, min_col=1, max_col=2):
            if row[0].value == vlan:
                vlan_name = row[1].value
                vlan_names.append(vlan_name)
                found = True
            break  # Exit inner loop once VLAN number is found
        if not found:
            vlan_names.append("")
    
    # Find the subnet base 
    base = ip[:ip.rfind('.') + 1] + '0'

    # Find the subnet mask 
    for row in ws_subnets.iter_rows(min_row=1, max_row=ws_subnets.max_row, min_col=1, max_col=16):
        if row[2].value == base:
            mask = row[3].value
            base_cidr = row[4].value
            base_cidr = base_cidr.replace('/','')
    
    # Setup which Vlan will be the disable "shutdown" 
    for row in ws_vlandatabase.iter_rows(min_row=1, max_row=ws_vlandatabase.max_row, min_col=1, max_col=16):
        row[2].value = str(row[2].value)
        if row[2].value.casefold() == "disabled":
            shutdown = row[0].value
            shutdown = str(shutdown)

    # Find Management IP 
    for row in ws_subnets.iter_rows(min_row=1, max_row=ws_subnets.max_row, min_col=1, max_col=16):
        row[0].value = str(row[0].value)
        if row[0].value.casefold() == "management network":
            management_subnet = row[2].value
            management_cidr = row[4].value
            management_cidr = management_cidr.replace('/','')

    # Find Client Server Networks
    client_networks = []
    client_cidrs = []
    for row in ws_subnets.iter_rows(min_row=1, max_row=ws_subnets.max_row, min_col=1, max_col=16):
        row[0].value = str(row[0].value)
        if row[0].value.casefold() == "client server network":
            client_network = row[2].value
            client_networks.append(client_network)
            client_cidr = row[4].value
            client_cidr = client_cidr.replace('/','')
            client_cidrs.append(client_cidr)

    # Find SNTP
    for row in ws_subnets.iter_rows(min_row=1, max_row=ws_subnets.max_row, min_col=1, max_col=16):
        if row[2].value == base:
            ntp1 = row[11].value
            ntp2 = row[12].value

    parts = [part.strip() for part in vlantags.split(',')]

    # Print to txt file (change the format if needed)
    if tag != prev_tag:
        with open(file_path, 'w') as f:
            f.write(f"! Copyright Inc\n")
            f.write(f"! {project_name}\n")
            f.write(f"! {customer}\n")
            f.write(f"! Cabinet Name {location}\n")
            f.write(f"! {network}\n")
            f.write(f"! Build Date: {build_date}\n\n")

            f.write(f"network dhcp config-load disable\n")
            f.write(f"network parms {ip} {mask} {gateway}\n\n")

            f.write(f"no network ipv6 operation\nnetwork ipv6 protocol none\n\n")

            f.write(f"network hidiscovery operation disable\n")
            f.write(f"network hidiscovery mode read-only\n\n")

            f.write(f"network management priority dot1p 7\nnetwork management access delete 1\nnetwork management access delete 1\n")

            f.write(f"network management access add 1 ip {base} mask {base_cidr} http disable https enable snmp enable telnet disable modbus-tcp disable ssh enable ethernet-ip disable profinet-io disable\nnetwork management access status 1 enable\nnetwork management access add 2 ip {management_subnet} mask {management_cidr} http disable https enable snmp enable telnet disable modbus-tcp disable ssh enable ethernet-ip disable profinet-io disable\nnetwork management access status 2 enable\nnetwork management access add 3 ip {client_networks[0]} mask {client_cidrs[0]} http disable https disable snmp enable telnet disable modbus-tcp disable ssh disable ethernet-ip disable profinet-io disable\nnetwork management access status 3 enable\nnetwork management access add 4 ip {client_networks[1]} mask {client_cidrs[1]} http disable https disable snmp enable telnet  disable modbus-tcp disable ssh disable ethernet-ip disable profinet-io disable\nnetwork management access status 4 enable\nnetwork management access operation\n\n")

            f.write(f"cli prompt {tag}\n")
            f.write(f"cli serial-timeout 5\n")
            f.write(f"cli banner operation enable\n\n")

            f.write(f"configure\n")
            f.write(f"system contact ABB_Service_Engineer\n")
            f.write(f"system location {location}\n")
            f.write(f"system name {tag}\n\n")

            f.write(f"config envm auto-update sd disable \n")
            f.write(f"config envm config-save sd disable \n")
            f.write(f"config envm sshkey-auto-update sd disable \n")
            f.write(f"config envm load-priority sd disable \n\n")

            f.write(f"clock summer-time mode usa\n\n")

            f.write(f"sntp client server add 1 {ntp1} port 123 description DC1\n")
            f.write(f"sntp client server add 2 {ntp2} port 123 description DC2\n")
            f.write(f"sntp client server mode 1 enable\n")
            f.write(f"sntp client server mode 2 enable\n")
            f.write(f"sntp client operation\n")

            f.write(f"passwords min-special-chars 0\n")
            f.write(f"users password admin {password_admin}\n")
            f.write(f"users add user\n")
            f.write(f"users password user {password_user}\n")
            f.write(f"users enable user\n\n")

            f.write(f"users password-policy-check admin enable\n")
            f.write(f"users password-policy-check user enable\n")
            f.write(f"passwords min-length 12\n")
            f.write(f"passwords min-special-chars 0\n\n")

            f.write(f"temperature upper-limit {max_temp}\n")
            f.write(f"temperature lower-limit {min_temp}\n\n")

            f.write(f"humidity upper-limit {max_humidity}\n")
            f.write(f"humidity lower-limit {min_humidity}\n\n")

            f.write(f"no telnet server\nno http server\nno snmp access version v1\nno snmp access version v2\nssh max-sessions 2\nssh timeout 5\n\n")

            f.write(f"https fingerprint-type sha256\nhttps certificate delete\nhttps certificate generate\n\n")

            f.write(f"cos-queue max-bandwidth 7 20\n\n")

            f.write(f"dos icmp-smurf-attack\ndos ip-land enable\ndos tcp-null\ndos tcp-syn-fin\ndos tcp-xmas\n\n")

            f.write(f"mac notification interval 30\nmac notification operation\n\n")

            f.write(f'system pre-login-banner operation\nsystem pre-login-banner text "***{network}***\\n\\nThis system is the property of {customer}\\n\\nOnly authorized personnel may use this system and all usage shall comply with the company security policy. Disconnect immediately if you do not fully agree on the company security policy or if the connection is unauthorized.\\n\\nUnauthorized use of this system is strictly prohibited and may be subject to criminal prosecution.\\n\\nAny usage of the system can be monitored and logged.\\n"\n\n')

            f.write(f"exit\nvlan database\n")
            for vlan_name, vlan in zip(vlan_names, vlans):
                f.write(f"vlan add {vlan}\n")
                f.write(f"name {vlan} {vlan_name}\n")
            f.write(f"exit\nconfigure\n\n")

    # This section writes the information for the individual ports
    with open(file_path, 'a') as f:
        f.write(f"interface 1/{port}\n")

        if vlantags == shutdown:
            f.write(f"shutdown\n")
        else:
            f.write(f"no shutdown\n")

        f.write(f"mtu {mtu}\nutilization alarm-threshold upper 7000\n")

        f.write(f"vlan participation auto 1\n")
        for part in parts:
            f.write(f"vlan participation include {part}\n")
            if tagged != "U":
                for part in parts:
                    f.write(f"vlan tagging {part}\n")

        f.write(f"vlan pvid {vlanpvid}\n")

        if rstp.casefold() == "admin edge port active":
            f.write(f"spanning-tree edge-port\n")
        elif rstp.casefold() == "rstp inactive":
            f.write(f"no spanning-tree mode\n")

        if speed.casefold() == "fdx-10":
            f.write(f"speed 10 full\n")
        elif speed.casefold() == "fdx-100":
            f.write(f"speed 100 full\n")
        elif speed.casefold() == "fdx-1000":
            f.write(f"speed 1000 full\n")
        elif speed.casefold() == "hdx-10":
            f.write(f"speed 10 half\n")
        elif speed.casefold() == "hdx-100":
            f.write(f"speed 100 half\n")
        elif speed.casefold() == "hdx-1000":
            f.write(f"speed 1000 half\n")
        f.write(f"exit\n\n")

    if tag != next_tag:
        with open(file_path, 'a') as f:

            # Subring Configuration 
            additions = 0
            title = 0
            role = None
            for row in range(4, ws.max_row + 1):
                cell_v = ws.cell(row=row, column=22).value
                cell_o = ws.cell(row=row, column=15).value
                cell_m = ws.cell(row=row, column=13).value
                cell_b = ws.cell(row=row, column=2).value
                if ("RSRM" in str(cell_m).upper()):
                    role = "redundant-manager"
                elif ("SRM" in str(cell_m).upper()):
                    role = "manager"
                if cell_v != ("N/A" or "NA" or "na" or "n/a" or ""):
                    if cell_b == tag:
                        if title == 0:
                            f.write("configure\n")
                            f.write("sub-ring operation\n")
                            title = 1
                        additions += 1
                        f.write(f'sub-ring add {additions} mode {role} vlan {cell_v} port 1/{cell_o} name "" mrp-domain 255.255.255.255.255.255.255.255.255.255.255.255.255.255.255.255\n')
            for i in range(additions): 
                f.write(f"sub-ring enable {i+1}\n")
            if additions != 0:
                f.write("exit\n\n")

            # Ring Configuration
            titles = 0
            roles = None
            for row in range(4, ws.max_row + 1):
                cell_u = ws.cell(row=row, column=21).value
                cell_o = ws.cell(row=row, column=15).value
                cell_m = ws.cell(row=row, column=13).value
                cell_b = ws.cell(row=row, column=2).value
                if ("MRM" in str(cell_m).upper()):
                    roles = "primary"
                elif ("MRC" in str(cell_m).upper()):
                    roles = "secondary"
                if cell_u != ("N/A" or "NA" or "na" or "n/a"):
                    if cell_b == tag:
                        if titles == 0:
                            f.write("configure\n")
                            f.write("mrp domain add default-domain\n")
                            titles = 1
                        f.write(f'mrp domain modify port {roles} port 1/{cell_o}\n')
                        if roles == "primary":
                            f.write(f"mrp domain modify mode manager")
                        if roles == "secondary":
                            f.write(f"mrp domain modify port secondary 1/{cell_o} fixed-backup disable\nmrp domain modify mode client\n")
            if titles != 0:
                f.write('mrp domain modify advanced-mode enable\nmrp domain modify manager-priority 32768\nmrp domain modify name ""\nmrp domain modify recovery-delay 200ms\nmrp domain modify {}\nmrp domain modify operation enable\nmrp operation enable\nexit\n\n')

            # Final Setup Output
            f.write(f"device-status monitor link-failure\ndevice-status monitor power-supply 1\ndevice-status monitor power-supply 2\ndevice-status monitor temperature\ndevice-status monitor humidity\ndevice-status trap enable\ndevice-status monitor ring-redundancy\n\n")

            f.write(f"signal-contact 1 mode manual\nsignal-contact 1 state close\nno signal-contact 1 monitor temperature\nno signal-contact 1 monitor power-supply 1\nno signal-contact 1 monitor power-supply 2\nno signal-contact 1 monitor humidity\n\n")

            f.write(f"security-status monitor extnvm-load-unsecure\nsecurity-status monitor extnvm-upd-enabled\nsecurity-status monitor pwd-change\nsecurity-status monitor pwd-min-length\nsecurity-status monitor pwd-policy-config\nsecurity-status monitor pwd-policy-inactive\n\n")

            f.write(f"security-status monitor hidisc-enabled\nsecurity-status monitor http-enabled\nsecurity-status monitor telnet-enabled\nsecurity-status monitor snmp-unsecure\n\n")

            f.write(f"no security-status monitor https-certificate\nno security-status monitor no-link-enabled\nno security-status monitor sysmon-enabled\n\n")

            f.write(f"security-status trap enable\n\n")

            f.write(f"spanning-tree bpdu-guard\nauto-disable reason bpdu-rate enable\ninterface all\nauto-disable timer 30\nexit\n\n")

            f.write(f"!Exit configure mode\nexit\n\n")
            f.write(f"network management vlan {network_management_vlan}\n")
            f.write(f"save\n\nreboot\n")
